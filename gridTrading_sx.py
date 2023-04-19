import os
import sys
import pyyjj
import kungfu.yijinjing.time as kft
from kungfu.wingchun.constants import *
import math
import kungfu.wingchun.utils as wc_utils
import openpyxl
from threading import Lock

source = Source.XTP
sz_exchange = Exchange.SZE
sh_exchange = Exchange.SSE

class Stock:
    def __init__(self,strStockCode,strExchange,fInitBasisPrice,fSellPriceDelta,
                    fBuyPriceDelta,fPriceUpperBound,fPriceLowerBound,iAmountPerEntrust,
                    iMaxBuyAmount,iMaxSellAmount,iMaxNettingAmount,index = -1):
        self.strStockCode = strStockCode         # 股票代码 例 000666
        self.strExchange = strExchange           # 交易所 例 SSE
        self.fInitBasisPrice = fInitBasisPrice   # 初始基准价格
        self.fSellPriceDelta = fSellPriceDelta   # 卖出价差，为百分比，如，0.02代表价差百分比为2%
        self.fBuyPriceDelta = fBuyPriceDelta     # 买入价差，为百分比
        self.fPriceUpperBound = fPriceUpperBound # 价格上限
        self.fPriceLowerBound = fPriceLowerBound # 价格下限
        self.iAmountPerEntrust = iAmountPerEntrust # 单次委托数量
        self.iMaxBuyAmount = iMaxBuyAmount       # 最大买入数量
        self.iMaxSellAmount = iMaxSellAmount     # 最大卖出数量
        self.iMaxNettingAmount = iMaxNettingAmount     # 最大轧差
        self.iBuyAmount = 0                      # 委托买入数量
        self.iSellAmount = 0                     # 委托卖出数量
        self.fCurrBasisPrice = fInitBasisPrice   # 储存上一次的买卖成交均价
        self.isSell = True
        self.isBuy = True
        self.index = index #用于记录在excel中的索引序号,快速保存该股票的最新交易价格


def read_excel(context, sz , sh ):
    wb = openpyxl.load_workbook(context.filename) # 读取xlsx文件
    sheet1 = wb.active
    context.account = str(sheet1.cell(1,2).value)
    #context.log.info("the account is {}, its type is {}".format(context.account,type(context.account)))    
    nrows = sheet1.max_row + 1 
    ncols = sheet1.max_column + 1
    for i in range(3,nrows):  #modified by shizhao on 20191120
        instrument_id = sheet1.cell(i,1).value
        exchange_id = sheet1.cell(i,2).value
        #context.log.info("type of instrument_id is {}; type of exchange_id is {};{}".format(instrument_id, exchange_id, isinstance(instrument_id,str)))
        if isinstance(instrument_id,str) and exchange_id == sz_exchange:
            sz.append(instrument_id)
        elif isinstance(instrument_id,str) and exchange_id == sh_exchange:
            sh.append(instrument_id)
        else:
            #added by shizhao on 20191120
            context.log.info("warning: error instrument_id or exchange_id info in the {}th row".format(i))
            pass

        if isinstance(instrument_id,str) and (exchange_id == sz_exchange or exchange_id == sh_exchange): #modified by shizhao on 20191125
            key = instrument_id + exchange_id
            context.stock_dict[key] = Stock(sheet1.cell(i,1).value, sheet1.cell(i,2).value, float(sheet1.cell(i,3).value), float(sheet1.cell(i,4).value)/100.0, float(sheet1.cell(i,5).value)/100.0, float(sheet1.cell(i,6).value), float(sheet1.cell(i,7).value), sheet1.cell(i,8).value, sheet1.cell(i,9).value, sheet1.cell(i,10).value ,sheet1.cell(i,11).value ,i)

def convert_time_nano(context , time_str):
    time_str = context.trading_day.strftime("%Y-%m-%d") + " " + time_str
    return context.strptime(time_str,"%Y-%m-%d %H:%M:%S")

def pre_start(context):
    context.log.info("pre run strategy")   
    context.stock_dict = {}
    sz = []
    sh = []
    context.account = ""                  #added by shizhao on 20191125
    context.customized_trading_time_begin = convert_time_nano(context,"09:30:00")   #modified by shizhao on 20191202
    context.customized_trading_time_end   = convert_time_nano(context,"22:56:59")   #modified by shizhao on 20191204    
    current_path = context.getParamFileDir()
    context.filename = os.path.join(current_path, "grid_target.xlsx")
    read_excel(context ,sz , sh )
    context.log.info("sz:{}.".format(sz))
    context.log.info("sh:{}.".format(sh))
    context.log.info(context.stock_dict.keys())
    context.log.info("customized beginning trading time is {}".format(kft.strftime(context.customized_trading_time_begin)))
    context.log.info("customized ending trading time is {}".format(kft.strftime(context.customized_trading_time_end)))        
    context.add_account(source, context.account, 100000000.0) 
    context.subscribe_market_data(source, sz , sz_exchange) #modified by shizhao on 20191120
    context.subscribe_market_data(source, sh , sh_exchange)


def pre_stop(context):
    context.log.info("pre stop strategy")


def on_quote(context, quote):
    #added by shizhao on 20191125
    
    if pyyjj.now_in_nano() <= context.customized_trading_time_begin or pyyjj.now_in_nano() >= context.customized_trading_time_end:
        context.log.info("warning:当前时间是{},不在自定义交易时间区间内！！！".format(kft.strftime(pyyjj.now_in_nano())))
        return
    
    key = quote.instrument_id + quote.exchange_id
    #context.log.info("instrument_id:{} last_price:{}".format(quote.instrument_id, quote.last_price))
    if key in context.stock_dict:
        stock = context.stock_dict[key]
        context.log.debug("instrument_id:{}, last_price:{}, bid_price[0]:{}, ask_price[0]:{}, fCurrBasisPrice:{}, upper_limit_price is {},lower_limit_price is {}".format(
                         quote.instrument_id, quote.last_price, quote.bid_price[0], quote.ask_price[0], stock.fCurrBasisPrice, quote.upper_limit_price, quote.lower_limit_price))
        if quote.last_price > stock.fCurrBasisPrice:#卖的可能
            # context.log.info("卖:instrument_id:{}               last_price:{}".format(quote.instrument_id, quote.last_price ))
            # context.log.info(quote.last_price - stock.fCurrBasisPrice)
            # context.log.info(stock.fCurrBasisPrice + stock.fSellPriceDelta <= stock.fPriceUpperBound)
            # context.log.info(stock.isSell)
            # context.log.info(stock.iSellAmount + stock.iAmountPerEntrust <= stock.iMaxSellAmount)
            # context.log.info(stock.iSellAmount+stock.iAmountPerEntrust-stock.iBuyAmount <= stock.iMaxNettingAmount)
            
            #added by shizhao on 20191130
            rate_of_price_increase = quote.last_price/stock.fCurrBasisPrice - 1.0
            multiple = int(rate_of_price_increase/stock.fSellPriceDelta)
            if 0 == multiple:
                return
            new_entrust_price = round(stock.fCurrBasisPrice*(1.0 + stock.fSellPriceDelta*multiple), 2)
            if new_entrust_price <= stock.fPriceUpperBound and new_entrust_price <= quote.upper_limit_price and new_entrust_price >= quote.lower_limit_price and stock.isSell and (stock.iSellAmount + stock.iAmountPerEntrust <= stock.iMaxSellAmount) and stock.iSellAmount+stock.iAmountPerEntrust-stock.iBuyAmount <= stock.iMaxNettingAmount:
                order_id = context.insert_order(quote.instrument_id, quote.exchange_id, context.account, new_entrust_price, int(stock.iAmountPerEntrust)*multiple , PriceType.Limit, Side.Sell, Offset.Open)
                context.log.info("insert_order({},{},{},{},{},{},{},{}), order_id is {}".format(quote.instrument_id, 
                                  quote.exchange_id, context.account, new_entrust_price, int(stock.iAmountPerEntrust)*multiple , 
                                  PriceType.Limit, Side.Sell, Offset.Open, order_id))  #modified by shizhao on 20191217
                context.log.info("last_price is {}, fCurrBasisPrice is {}, new_entrust_price is {}, fPriceUpperBound is {}, upper_limit_price is {}, lower_limit_price is {}, iSellAmount is {}, iAmountPerEntrust is {}, iMaxSellAmount is {}, iBuyAmount is {}, iMaxNettingAmount is {}".format(
                                 quote.last_price, stock.fCurrBasisPrice, new_entrust_price, stock.fPriceUpperBound, quote.upper_limit_price, 
                                 quote.lower_limit_price, stock.iSellAmount, stock.iAmountPerEntrust, stock.iMaxSellAmount, 
                                 stock.iBuyAmount, stock.iMaxNettingAmount))  #modified by shizhao on 20191217
                if order_id > 0:###已经考虑到报单错误或者交易所拒单的情况
                    stock.iSellAmount = stock.iSellAmount + stock.iAmountPerEntrust*multiple
                    stock.isSell = False
        
        elif quote.last_price < stock.fCurrBasisPrice:#买的可能
            # context.log.info("买:instrument_id:{}               last_price:{}".format(quote.instrument_id, quote.last_price ))
            # context.log.info(stock.fCurrBasisPrice -quote.last_price >= stock.fBuyPriceDelta)
            # context.log.info(stock.fCurrBasisPrice - stock.fBuyPriceDelta >= stock.fPriceLowerBound)
            # context.log.info(stock.isBuy)
            # context.log.info(stock.iBuyAmount + stock.iAmountPerEntrust <= stock.iMaxBuyAmount)
            # context.log.info(stock.iBuyAmount+stock.iAmountPerEntrust-stock.iSellAmount <= stock.iMaxNettingAmount)
            
            #added by shizhao on 20191130
            rate_of_price_decrease = 1.0 - quote.last_price/stock.fCurrBasisPrice
            multiple = int(rate_of_price_decrease/stock.fBuyPriceDelta)
            if 0 == multiple:
                return            
            new_entrust_price = round(stock.fCurrBasisPrice*(1.0 - stock.fBuyPriceDelta*multiple), 2)
            if new_entrust_price >= stock.fPriceLowerBound and new_entrust_price >= quote.lower_limit_price and new_entrust_price <= quote.upper_limit_price and stock.isBuy and (stock.iBuyAmount + stock.iAmountPerEntrust <= stock.iMaxBuyAmount) and stock.iBuyAmount+stock.iAmountPerEntrust-stock.iSellAmount <= stock.iMaxNettingAmount:
                order_id = context.insert_order(quote.instrument_id, quote.exchange_id, context.account, new_entrust_price, int(stock.iAmountPerEntrust)*multiple , PriceType.Limit, Side.Buy, Offset.Open)
                context.log.info("insert_order({},{},{},{},{},{},{},{}), order_id is {}".format(quote.instrument_id, 
                                  quote.exchange_id, context.account, new_entrust_price, int(stock.iAmountPerEntrust)*multiple , 
                                  PriceType.Limit, Side.Buy, Offset.Open, order_id)) #modified by shizhao on 20191217
                context.log.info("last_price is {}, fCurrBasisPrice is {}, new_entrust_price is {}, fPriceLowerBound is {}, lower_limit_price is {}, upper_limit_price is {}, iBuyAmount is {}, iAmountPerEntrust is {}, iMaxBuyAmount, iSellAmount is {}, iMaxNettingAmount is {}".format(
                                 quote.last_price, stock.fCurrBasisPrice, new_entrust_price, stock.fPriceLowerBound, 
                                 quote.lower_limit_price, quote.upper_limit_price, stock.iBuyAmount, stock.iAmountPerEntrust, 
                                 stock.iMaxBuyAmount, stock.iSellAmount, stock.iMaxNettingAmount)) #modified by shizhao on 20191217
                if order_id > 0:
                    stock.iBuyAmount = stock.iBuyAmount + stock.iAmountPerEntrust*multiple
                    stock.isBuy = False
        else:#no need any operation
            pass


def on_transaction(context, transaction):
    #context.log.info("{} {}".format(transaction.instrument_id, transaction.exchange_id))
    pass

def on_entrust(context, entrust):
    #context.log.info("{} {}".format(entrust.instrument_id, entrust.exchange_id))
    pass

def on_order(context, order):
    context.log.info('order received: [instrument_id]{} [volume]{} [price]{}'.format(order.instrument_id, order.volume, order.limit_price)) #added by shizhao on 20191217
    key = order.instrument_id + order.exchange_id
    if key in context.stock_dict:
        stock = context.stock_dict[key]
        if wc_utils.is_final_status(order.status):
            if order.side == Side.Buy:
                stock.iBuyAmount = stock.iBuyAmount - order.volume_left
            elif order.side == Side.Sell:
                stock.iSellAmount = stock.iSellAmount - order.volume_left
            else:
                pass
                    
            if order.volume_traded > 0:#有成交股数
                wb = openpyxl.load_workbook(context.filename) # 读取xlsx文件
                ws = wb.active
                stock.fCurrBasisPrice = order.amount_traded/order.volume_traded #修改最新成交价
                ws.cell(stock.index,3).value = stock.fCurrBasisPrice
                wb.save(context.filename)
                context.log.info("save new price because the final_status order,order_id is {}".format(order.order_id)) #modified by shizhao on 20191217
                
                stock.isSell = True                
                stock.isBuy = True
            else:#无成交股数
                if order.side == Side.Buy:#需要修改下
                    stock.isBuy = True
                else:#取消卖出
                    stock.isSell = True                    